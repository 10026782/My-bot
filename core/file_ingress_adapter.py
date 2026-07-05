# core/file_ingress_adapter.py — C90: Structured File Capture ingress adapter
#
# "File upload is a new INGRESS SOURCE ADAPTER only. It is NOT a new capture
# pipeline, NOT a new classifier, NOT a new write path." (C90 spec)
#
# Parses xlsx/csv bytes into a list of per-row text strings, one per data
# row (the header row is consumed as field labels, not emitted as its own
# event). Each returned row-text is meant to be fed, one at a time, through
# the EXISTING classify_ingress()/handle_lead_candidate() pipeline exactly
# like a text message would be — no lead extraction, no tier logic, no
# Airtable/Gateway calls happen in this module.

from __future__ import annotations

import csv
import io
import logging

from core.ingress_envelope import IngressEnvelope

logger = logging.getLogger(__name__)

# C94 Stage ב — channel this adapter's files currently arrive on → provider.
# File upload today only reaches this adapter via Telegram document messages
# (see app.py's _handle_telegram_media); extend when another channel gains
# file upload.
_CHANNEL_TO_PROVIDER = {
    "telegram": "telegram_bot_api",
}


class FileParseError(Exception):
    """Raised when the file itself can't be parsed at all (bad format, corrupt, empty)."""


def parse_structured_file_rows(file_bytes: bytes, filename: str) -> list[str]:
    """
    מחזיר list של מחרוזות טקסט, אחת לכל שורת נתונים (לא כולל שורת כותרות).
    כל שורה מיוצגת כ-"col1: val1 | col2: val2 | ..." — צורה שממשיכה לעבוד
    עם ה-extraction הקיים (regex על טלפון/שם), בלי תלות במיקום עמודות.

    "no merging/dropping": כל שורת נתונים מיוצרת כאירוע נפרד — אין קאפ/חיתוך
    כאן (מגבלת בטיחות תפעולית, אם נדרשת, היא באחריות הקורא ולא של הפרסר).
    שורה ריקה לגמרי (כל התאים None/ריק) מדולגת — אין בה מידע כלל, לא "נשמטת".
    שורה שנכשלת לפרסר (תא לא ניתן ל-stringify וכו') לא נעלמת בשקט — מוחזרת
    כטקסט raw מסומן ("[malformed_row] ...") כדי ש-classify_ingress() עדיין
    ירוץ עליה ותמיד תיווצר עבורה תצפית — Tier 4/5 בטוח, לעולם לא dropped.

    זורק FileParseError אם הקובץ כולו לא ניתן לפתיחה (פורמט שגוי/corrupt/ריק).
    """
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _read_csv_rows(file_bytes)
    elif lower.endswith(".xlsx"):
        rows = _read_xlsx_rows(file_bytes)
    else:
        raise FileParseError(f"unsupported file extension: {filename}")

    if not rows:
        raise FileParseError("file contains no rows")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = rows[1:]

    row_texts: list[str] = []
    for row in data_rows:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # fully empty row — zero information, not "dropped" content
        try:
            row_texts.append(_row_to_text(header, row))
        except Exception as exc:
            logger.warning("[FileIngress] malformed row in %s: %s", filename, exc)
            raw = " ".join(str(c) for c in row if c is not None)
            row_texts.append(f"[malformed_row] {raw}")

    return row_texts


# C94 Stage ב — this adapter's ONLY C94 responsibility: turn one already-
# parsed row into a valid IngressEnvelope. It does not call classify_ingress()
# itself, does not know about tiers/ActionGateway, and never touches
# EvidenceTrace (that's a byproduct the caller builds AFTER running the row
# through the existing, unchanged C89/C90 pipeline — see app.py).
def build_file_row_envelope(
    *, identity, channel: str, filename: str, row_index: int, row_text: str,
) -> IngressEnvelope:
    """
    בונה IngressEnvelope לשורת קובץ בודדת, לפני כל כניסה ל-classify_ingress().
    provider נגזר מה-channel שממנו הועלה הקובץ (_CHANNEL_TO_PROVIDER) — לא
    מנוחש/הארדקוד בקריאה עצמה, כדי שהחלפת provider עתידית (אם יתווסף ערוץ
    file upload נוסף) תדרוש רק עדכון המיפוי הזה.

    source_ref (לא raw_ref — ראה C94-A.1) הוא מצביע קבוע-מראש ("file:<שם
    קובץ>:row<אינדקס>"), זמין תמיד לפני סיווג, ואינו תלוי ב-C89's
    raw-capture mechanism. envelope_id מזוהה אוטומטית (uuid) על ידי
    IngressEnvelope עצמו.
    """
    provider = _CHANNEL_TO_PROVIDER.get(channel, channel)
    return IngressEnvelope(
        source_channel="file_upload",
        provider=provider,
        raw_event_id=f"{filename}:row{row_index}",
        sender_identity=identity.memory_key,
        normalized_text=row_text,
        attachments=(filename,),
        source_ref=f"file:{filename}:row{row_index}",
    )


def _row_to_text(header: list[str], row: tuple) -> str:
    # NOTE: ", " is deliberate, not cosmetic — a " | " separator would make
    # any row with 3+ populated columns collide with classify_ingress()'s
    # existing pipe-table Tier-4 marker (_TABLE_RE: "2+ pipe-separated
    # fields"), force-classifying every multi-column row as Tier 4 as a
    # side effect of formatting rather than genuine ambiguity. ", " reads
    # as an ordinary comma-separated lead line (what Tier 1-3 extraction
    # already expects) and doesn't collide with any Tier-4 hard marker.
    parts = []
    for i, cell in enumerate(row):
        label = header[i] if i < len(header) and header[i] else f"col{i + 1}"
        value = "" if cell is None else str(cell).strip()
        if value:
            parts.append(f"{label}: {value}")
    return ", ".join(parts)


def _read_csv_rows(file_bytes: bytes) -> list[tuple]:
    try:
        text = file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise FileParseError(f"csv decode error: {exc}") from exc

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
    except csv.Error:
        dialect = csv.excel

    try:
        return [tuple(r) for r in csv.reader(io.StringIO(text), dialect)]
    except csv.Error as exc:
        raise FileParseError(f"csv parse error: {exc}") from exc


def _read_xlsx_rows(file_bytes: bytes) -> list[tuple]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise FileParseError("openpyxl not installed") from exc

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise FileParseError(f"xlsx parse error: {exc}") from exc

    if not workbook.worksheets:
        raise FileParseError("xlsx contains no sheets")

    sheet = workbook.worksheets[0]
    return [tuple(r) for r in sheet.iter_rows(values_only=True)]
