from __future__ import annotations

import csv
from pathlib import Path

from ..result import failure, success


def convert(source: Path, output: Path, input_type: str, output_type: str):
    if input_type == "csv" and output_type == "xlsx":
        return _csv_to_xlsx(source, output)
    if input_type == "xlsx" and output_type == "csv":
        return _xlsx_to_csv(source, output)
    return failure(f"unsupported table conversion: {input_type} -> {output_type}")


def _csv_to_xlsx(source: Path, output: Path):
    from openpyxl import Workbook

    rows = _read_csv(source)
    if not rows:
        return failure("csv contains no rows")
    width = len(rows[0])
    if any(len(row) != width for row in rows):
        return failure("ragged csv rows are not supported")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    workbook.save(str(output))
    return success(str(output))


def _xlsx_to_csv(source: Path, output: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(str(source), data_only=False)
    non_empty_sheets = [
        sheet
        for sheet in workbook.worksheets
        if any(any(cell.value is not None for cell in row) for row in sheet.iter_rows())
    ]
    if len(non_empty_sheets) != 1:
        return failure("xlsx conversion supports exactly one non-empty sheet")
    sheet = non_empty_sheets[0]
    if sheet.merged_cells.ranges:
        return failure("xlsx files with merged cells are not supported")
    if getattr(sheet, "_charts", None) or getattr(sheet, "_images", None):
        return failure("xlsx files with charts or images are not supported")

    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if value is None else value for value in row])
    return success(str(output))


def _read_csv(source: Path) -> list[list[str]]:
    with source.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
        return list(csv.reader(handle, dialect))
